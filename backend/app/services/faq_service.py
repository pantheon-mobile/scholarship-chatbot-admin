from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile, ZipFile, is_zipfile
from zoneinfo import ZoneInfo

from fastapi import UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException

from app.models.faq import Faq
from app.repositories.faq import FaqRepository
from app.schemas.faq import (
    FaqBulkDeleteRequest,
    FaqClassificationResponse,
    FaqCreateRequest,
    FaqDetailResponse,
    FaqFilters,
    FaqListResponse,
    FaqImportResponse,
    FaqImportRowError,
    FaqResponse,
    FaqSimilarQuestionResponse,
    FaqUpdateRequest,
)


FAQ_IMPORT_MAX_BYTES = 10 * 1024 * 1024
FAQ_IMPORT_MAX_ROWS = 1000
FAQ_IMPORT_SIMILAR_COUNT = 10
FAQ_IMPORT_FIXED_HEADERS = [
    "ID", "質問", "回答",
    *[f"類似質問{index}" for index in range(1, FAQ_IMPORT_SIMILAR_COUNT + 1)],
]


@dataclass
class FaqImportEntry:
    row_number: int
    faq_id: int | None
    question: str
    answer: str
    similar_questions: list[str]
    classifications: list[tuple[int, int]]
    chat_enabled: bool


class FaqError(Exception):
    def __init__(self, code: str, message: str, errors: list[FaqImportRowError] | None = None) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.errors = errors


class FaqService:
    def __init__(self, repository: FaqRepository) -> None:
        self.repository = repository

    async def validate_filters(self, filters: FaqFilters) -> dict[str, int]:
        type_ids: dict[str, int] = {}
        for index in range(1, 5):
            type_code = f"FAQ_TYPE_{index}"
            value_id = getattr(filters, f"classification_{index}_value_id")
            if value_id is not None:
                type_id = await self.repository.resolve_value_type(type_code, value_id)
                if type_id is None:
                    raise FaqError("INVALID_FAQ_CLASSIFICATION", "指定された区分値が正しくありません。")
                type_ids[type_code] = type_id
        return type_ids

    @staticmethod
    def serialize(row: Faq) -> FaqResponse:
        assignments = sorted(row.classification_assignments, key=lambda item: item.classification_type.display_order)
        return FaqResponse(
            id=row.id,
            question=row.question,
            answer=row.answer,
            chat_enabled=row.chat_enabled,
            updated_at=row.updated_at,
            version=row.version,
            classifications=[FaqClassificationResponse(
                type_code=item.classification_type.type_code,
                classification_type_id=item.classification_type_id,
                classification_value_id=item.classification_value_id,
                display_label=item.classification_type.display_label,
                value_name=item.classification_value.value_name,
            ) for item in assignments],
        )

    @classmethod
    def serialize_detail(cls, row: Faq) -> FaqDetailResponse:
        base = cls.serialize(row)
        return FaqDetailResponse(
            **base.model_dump(),
            created_at=row.created_at,
            similar_questions=[FaqSimilarQuestionResponse(
                id=item.id, question=item.question, display_order=item.display_order,
            ) for item in sorted(row.similar_questions, key=lambda item: item.display_order)],
        )

    @staticmethod
    def normalize_required(value: str, *, required_code: str, required_message: str, max_length: int, long_code: str, long_message: str) -> str:
        normalized = value.strip()
        if not normalized:
            raise FaqError(required_code, required_message)
        if len(normalized) > max_length:
            raise FaqError(long_code, long_message)
        return normalized

    @classmethod
    def normalize_input_text(cls, payload: FaqCreateRequest) -> tuple[str, str, list[str]]:
        question = cls.normalize_required(
            payload.question, required_code="FAQ_QUESTION_REQUIRED", required_message="質問を入力してください。",
            max_length=500, long_code="FAQ_QUESTION_TOO_LONG", long_message="質問は500文字以内で入力してください。",
        )
        answer = cls.normalize_required(
            payload.answer, required_code="FAQ_ANSWER_REQUIRED", required_message="回答を入力してください。",
            max_length=1000, long_code="FAQ_ANSWER_TOO_LONG", long_message="回答は1000文字以内で入力してください。",
        )
        similar_questions = [cls.normalize_required(
            value, required_code="FAQ_SIMILAR_QUESTION_REQUIRED", required_message="類似質問を入力してください。",
            max_length=500, long_code="FAQ_SIMILAR_QUESTION_TOO_LONG", long_message="類似質問は500文字以内で入力してください。",
        ) for value in payload.similar_questions]
        return question, answer, similar_questions

    async def resolve_create_classifications(self, payload: FaqCreateRequest) -> list[tuple[int, int]]:
        assignments: list[tuple[int, int]] = []
        for index in range(1, 5):
            value_id = getattr(payload, f"classification_{index}_value_id")
            if value_id is None:
                continue
            resolved = await self.repository.get_value_type(value_id)
            if resolved is None:
                raise FaqError("FAQ_CLASSIFICATION_NOT_FOUND", "指定された区分値が見つかりません。")
            type_id, type_code = resolved
            if type_code != f"FAQ_TYPE_{index}":
                raise FaqError("INVALID_FAQ_CLASSIFICATION", "指定された区分値が正しくありません。")
            assignments.append((type_id, value_id))
        return assignments

    async def get_detail(self, faq_id: int) -> FaqDetailResponse:
        row = await self.repository.get_detail(faq_id)
        if row is None:
            raise FaqError("FAQ_NOT_FOUND", "指定されたFAQが見つかりません。")
        return self.serialize_detail(row)

    async def create(self, payload: FaqCreateRequest) -> FaqDetailResponse:
        try:
            question, answer, similar_questions = self.normalize_input_text(payload)
            classifications = await self.resolve_create_classifications(payload)
            faq_id = await self.repository.create(
                question=question, answer=answer, similar_questions=similar_questions,
                classifications=classifications, chat_enabled=payload.chat_enabled,
            )
            await self.repository.commit()
            return await self.get_detail(faq_id)
        except FaqError:
            await self.repository.rollback()
            raise
        except Exception:
            await self.repository.rollback()
            raise

    async def update(self, faq_id: int, payload: FaqUpdateRequest) -> FaqDetailResponse:
        row = await self.repository.get_detail(faq_id)
        if row is None:
            raise FaqError("FAQ_NOT_FOUND", "指定されたFAQが見つかりません。")
        if row.version != payload.version:
            raise FaqError("FAQ_VERSION_CONFLICT", "他の操作で情報が更新されています。再読み込みしてください。")

        try:
            question, answer, similar_questions = self.normalize_input_text(payload)
            classifications = await self.resolve_create_classifications(payload)
            updated = await self.repository.update(
                faq_id,
                version=payload.version,
                question=question,
                answer=answer,
                similar_questions=similar_questions,
                classifications=classifications,
                chat_enabled=payload.chat_enabled,
            )
            if not updated:
                raise FaqError("FAQ_VERSION_CONFLICT", "他の操作で情報が更新されています。再読み込みしてください。")
            await self.repository.commit()
            return await self.get_detail(faq_id)
        except FaqError:
            await self.repository.rollback()
            raise
        except Exception as error:
            await self.repository.rollback()
            raise FaqError("FAQ_UPDATE_FAILED", "FAQの更新に失敗しました。") from error

    async def list(self, filters: FaqFilters) -> FaqListResponse:
        type_ids = await self.validate_filters(filters)
        rows, total_count, total_pages = await self.repository.list(filters, type_ids)
        if total_count and filters.page > total_pages:
            raise FaqError("PAGE_NOT_FOUND", "ページがありません。")
        return FaqListResponse(
            items=[self.serialize(row) for row in rows], page=filters.page, page_size=filters.page_size,
            total_count=total_count, total_pages=total_pages, sort=filters.sort, order=filters.order,
        )

    async def delete(self, faq_id: int, version: int) -> None:
        row = await self.repository.get(faq_id)
        if row is None:
            raise FaqError("FAQ_NOT_FOUND", "指定されたFAQが見つかりません。")
        if not await self.repository.delete_one(faq_id, version):
            raise FaqError("FAQ_VERSION_CONFLICT", "他の操作で情報が更新されています。再読み込みしてください。")

    async def bulk_delete(self, payload: FaqBulkDeleteRequest) -> int:
        try:
            return await self.repository.bulk_delete(payload.items)
        except LookupError:
            raise FaqError("FAQ_NOT_FOUND", "削除対象のFAQが見つかりません。") from None
        except ValueError as error:
            if str(error) == "version_mismatch":
                raise FaqError("FAQ_VERSION_CONFLICT", "他の操作で情報が更新されています。再読み込みしてください。") from None
            raise FaqError("INVALID_FAQ_DELETE_TARGETS", "削除対象が不正です。") from None

    async def export_excel(self, filters: FaqFilters, labels: dict[str, str]) -> bytes:
        type_ids = await self.validate_filters(filters)
        rows, _, _ = await self.repository.list(filters, type_ids, paginate=False)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "FAQ一覧"
        worksheet.append(["ID", "質問", "回答", *[labels.get(f"FAQ_TYPE_{i}", f"区分{i}") for i in range(1, 5)], "チャット利用", "更新日時"])
        jst = ZoneInfo("Asia/Tokyo")
        for row in rows:
            values = {item.classification_type.type_code: item.classification_value.value_name for item in row.classification_assignments}
            worksheet.append([
                row.id, row.question, row.answer, *[values.get(f"FAQ_TYPE_{i}", "") for i in range(1, 5)],
                "公開" if row.chat_enabled else "非公開", row.updated_at.astimezone(jst).strftime("%Y/%m/%d %H:%M"),
            ])
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    async def create_import_template(self) -> bytes:
        labels = await self.repository.list_type_labels()
        headers = [
            *FAQ_IMPORT_FIXED_HEADERS,
            *[labels.get(f"FAQ_TYPE_{index}", f"区分{index}") for index in range(1, 5)],
            "チャット利用",
        ]
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "FAQ一括登録更新"
        worksheet.append(headers)
        worksheet.freeze_panes = "A2"
        fill = PatternFill(fill_type="solid", fgColor="3D5AFE")
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
        widths = [12, 42, 56, *([30] * FAQ_IMPORT_SIMILAR_COUNT), *([22] * 4), 16]
        for index, width in enumerate(widths, start=1):
            worksheet.column_dimensions[chr(64 + index) if index <= 26 else "A"].width = width
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    async def _read_import_file(upload: UploadFile) -> bytes:
        filename = upload.filename or ""
        if Path(filename).suffix.lower() != ".xlsx":
            raise FaqError("FAQ_IMPORT_INVALID_FORMAT", "xlsx形式のファイルを選択してください。")
        content = await upload.read(FAQ_IMPORT_MAX_BYTES + 1)
        if len(content) > FAQ_IMPORT_MAX_BYTES:
            raise FaqError("FAQ_IMPORT_FILE_TOO_LARGE", "ファイルサイズは10MB以下にしてください。")
        if not content or not is_zipfile(BytesIO(content)):
            raise FaqError("FAQ_IMPORT_INVALID_FORMAT", "有効なxlsxファイルを選択してください。")
        try:
            with ZipFile(BytesIO(content)) as archive:
                names = {name.lower() for name in archive.namelist()}
                content_types = archive.read("[Content_Types].xml").lower()
                if any(name.endswith("vbaproject.bin") for name in names) or b"macroenabled" in content_types:
                    raise FaqError("FAQ_IMPORT_INVALID_FORMAT", "マクロを含むExcelファイルは使用できません。")
        except (BadZipFile, KeyError):
            raise FaqError("FAQ_IMPORT_INVALID_FORMAT", "有効なxlsxファイルを選択してください。") from None
        return content

    @staticmethod
    def _error(row: int, column: str, code: str, message: str) -> FaqImportRowError:
        return FaqImportRowError(row=row, column=column, code=code, message=message)

    @staticmethod
    def _text(value: object) -> str:
        return "" if value is None else str(value).strip()

    @classmethod
    def _parse_id(cls, value: object) -> int | None:
        if value is None or (isinstance(value, str) and not value.strip()):
            return None
        if isinstance(value, bool):
            raise ValueError
        if isinstance(value, int) and value >= 1:
            return value
        if isinstance(value, float) and value.is_integer() and value >= 1:
            return int(value)
        if isinstance(value, str) and value.strip().isdigit() and int(value.strip()) >= 1:
            return int(value.strip())
        raise ValueError

    async def import_excel(self, upload: UploadFile) -> FaqImportResponse:
        content = await self._read_import_file(upload)
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=False, keep_links=False)
        except (InvalidFileException, BadZipFile, KeyError, OSError, ValueError):
            raise FaqError("FAQ_IMPORT_INVALID_FORMAT", "有効なxlsxファイルを選択してください。") from None

        worksheet = workbook.active
        expected_column_count = len(FAQ_IMPORT_FIXED_HEADERS) + 5
        if worksheet.max_column != expected_column_count:
            workbook.close()
            raise FaqError("FAQ_IMPORT_INVALID_COLUMNS", "Excelの列数または列順が正しくありません。")
        header_cells = next(worksheet.iter_rows(min_row=1, max_row=1, max_col=expected_column_count))
        headers = [self._text(cell.value) for cell in header_cells]
        if (
            any(cell.data_type == "f" for cell in header_cells)
            or headers[:len(FAQ_IMPORT_FIXED_HEADERS)] != FAQ_IMPORT_FIXED_HEADERS
            or any(not value for value in headers[len(FAQ_IMPORT_FIXED_HEADERS):len(FAQ_IMPORT_FIXED_HEADERS) + 4])
            or headers[-1] != "チャット利用"
        ):
            workbook.close()
            raise FaqError("FAQ_IMPORT_INVALID_COLUMNS", "Excelの列数または列順が正しくありません。")
        if worksheet.max_row - 1 > FAQ_IMPORT_MAX_ROWS:
            workbook.close()
            raise FaqError("FAQ_IMPORT_TOO_MANY_ROWS", "データ行は1000行以内にしてください。")

        classification_types = await self.repository.list_import_classifications()
        type_maps = {
            item.type_code: (item.id, {value.value_name: value.id for value in item.values})
            for item in classification_types
        }
        errors: list[FaqImportRowError] = []
        entries: list[FaqImportEntry] = []

        for row_number, cells in enumerate(worksheet.iter_rows(min_row=2, max_col=expected_column_count), start=2):
            if all(self._text(cell.value) == "" for cell in cells):
                continue
            formula_columns = {index for index, cell in enumerate(cells) if cell.data_type == "f"}
            for index in sorted(formula_columns):
                errors.append(self._error(row_number, headers[index], "FAQ_IMPORT_FORMULA_NOT_ALLOWED", "数式は入力できません。"))

            faq_id = None
            if 0 not in formula_columns:
                try:
                    faq_id = self._parse_id(cells[0].value)
                except ValueError:
                    errors.append(self._error(row_number, "ID", "FAQ_ID_INVALID", "IDは正の整数で入力してください。"))

            question = self._text(cells[1].value) if 1 not in formula_columns else ""
            answer = self._text(cells[2].value) if 2 not in formula_columns else ""
            for value, column, required_code, required_message, limit, long_code, long_message in [
                (question, "質問", "FAQ_QUESTION_REQUIRED", "質問を入力してください。", 500, "FAQ_QUESTION_TOO_LONG", "質問は500文字以内で入力してください。"),
                (answer, "回答", "FAQ_ANSWER_REQUIRED", "回答を入力してください。", 1000, "FAQ_ANSWER_TOO_LONG", "回答は1000文字以内で入力してください。"),
            ]:
                if not value and headers.index(column) not in formula_columns:
                    errors.append(self._error(row_number, column, required_code, required_message))
                elif len(value) > limit:
                    errors.append(self._error(row_number, column, long_code, long_message))

            similar_questions: list[str] = []
            for offset in range(FAQ_IMPORT_SIMILAR_COUNT):
                index = 3 + offset
                if index in formula_columns:
                    continue
                value = self._text(cells[index].value)
                if not value:
                    continue
                if len(value) > 500:
                    errors.append(self._error(row_number, headers[index], "FAQ_SIMILAR_QUESTION_TOO_LONG", "類似質問は500文字以内で入力してください。"))
                else:
                    similar_questions.append(value)

            classifications: list[tuple[int, int]] = []
            for offset in range(4):
                index = len(FAQ_IMPORT_FIXED_HEADERS) + offset
                if index in formula_columns:
                    continue
                value_name = self._text(cells[index].value)
                if not value_name:
                    continue
                type_code = f"FAQ_TYPE_{offset + 1}"
                type_definition = type_maps.get(type_code)
                value_id = type_definition[1].get(value_name) if type_definition else None
                if value_id is None:
                    errors.append(self._error(row_number, headers[index], "FAQ_CLASSIFICATION_NOT_FOUND", "指定された区分値が存在しません。"))
                else:
                    classifications.append((int(type_definition[0]), int(value_id)))

            chat_index = expected_column_count - 1
            chat_text = self._text(cells[chat_index].value) if chat_index not in formula_columns else ""
            if chat_index not in formula_columns and chat_text not in ("公開", "非公開"):
                errors.append(self._error(row_number, "チャット利用", "FAQ_CHAT_ENABLED_INVALID", "チャット利用は「公開」または「非公開」で入力してください。"))

            entries.append(FaqImportEntry(
                row_number=row_number, faq_id=faq_id, question=question, answer=answer,
                similar_questions=similar_questions, classifications=classifications,
                chat_enabled=chat_text == "公開",
            ))
        workbook.close()

        if not entries:
            await self.repository.rollback()
            raise FaqError("FAQ_IMPORT_EMPTY", "登録・更新するFAQがありません。")

        by_id: dict[int, list[FaqImportEntry]] = {}
        for entry in entries:
            if entry.faq_id is not None:
                by_id.setdefault(entry.faq_id, []).append(entry)
        for faq_id, duplicate_entries in by_id.items():
            if len(duplicate_entries) > 1:
                for entry in duplicate_entries:
                    errors.append(self._error(entry.row_number, "ID", "FAQ_ID_DUPLICATE", f"ID {faq_id} がExcel内で重複しています。"))

        existing_rows = await self.repository.get_for_update_many(list(by_id))
        existing = {int(row.id): row for row in existing_rows}
        for faq_id, id_entries in by_id.items():
            if faq_id not in existing:
                for entry in id_entries:
                    errors.append(self._error(entry.row_number, "ID", "FAQ_NOT_FOUND", "指定されたFAQが見つかりません。"))

        if errors:
            await self.repository.rollback()
            raise FaqError("FAQ_IMPORT_VALIDATION_ERROR", "入力内容にエラーがあります。", errors)

        created_count = 0
        updated_count = 0
        try:
            for entry in entries:
                values = dict(
                    question=entry.question, answer=entry.answer,
                    similar_questions=entry.similar_questions,
                    classifications=entry.classifications, chat_enabled=entry.chat_enabled,
                )
                if entry.faq_id is None:
                    await self.repository.create(**values)
                    created_count += 1
                else:
                    updated = await self.repository.update(entry.faq_id, version=int(existing[entry.faq_id].version), **values)
                    if not updated:
                        raise FaqError("FAQ_IMPORT_FAILED", "FAQの一括登録／更新に失敗しました。")
                    updated_count += 1
            await self.repository.commit()
        except FaqError:
            await self.repository.rollback()
            raise
        except Exception as error:
            await self.repository.rollback()
            raise FaqError("FAQ_IMPORT_FAILED", "FAQの一括登録／更新に失敗しました。") from error
        return FaqImportResponse(
            created_count=created_count,
            updated_count=updated_count,
            processed_count=created_count + updated_count,
        )
