import csv
from datetime import date, datetime
from io import BytesIO, StringIO

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.v1.auth import require_authenticated_session, require_system_admin_session
from app.core.db import get_db
from app.models.auth import AuthSession
from app.repositories.reporting import ReportingRepository
from app.schemas.reporting import ChatHistoryResponse
from app.services.reporting_service import ReportingError, ReportingService, utc_period


router = APIRouter(tags=["reporting"])


def operation_description(method: str, path: str) -> str:
    if path.endswith("/ingestion/run"):
        return "データ取り込み処理を今すぐ実行"
    resources = [
        ("/faq-classifications", "FAQ区分"),
        ("/faqs", "FAQ"),
        ("/data-source-types", "データソース区分"),
        ("/data-sources", "データソース"),
        ("/categories", "カテゴリ"),
        ("/usage/users.csv", "ユーザーリスト"),
        ("/usage/access-logs.csv", "アクセスログ"),
        ("/usage/operation-logs.csv", "操作ログ"),
    ]
    resource = next((label for prefix, label in resources if path.startswith(f"/api/v1{prefix}")), "管理データ")
    if path.endswith(".csv") or path.endswith("/export"):
        return f"{resource}をダウンロード"
    if path.endswith("/import"):
        return f"{resource}を一括登録・更新"
    if path.endswith("/bulk-delete"):
        return f"{resource}を一括削除"
    if path.endswith("/order"):
        return f"{resource}の表示順を変更"
    action = {"POST": "登録", "PUT": "更新", "PATCH": "更新", "DELETE": "削除"}.get(method, "操作")
    return f"{resource}を{action}"


def operation_kind(method: str, path: str) -> str:
    if path.endswith(".csv") or path.endswith("/export"):
        return "DOWNLOAD"
    if path.endswith("/import") or "/files" in path:
        return "UPLOAD"
    return {"POST": "CREATE", "PUT": "UPDATE", "PATCH": "UPDATE", "DELETE": "DELETE"}.get(method, "OTHER")


def parsed_user_ids(value: str | None) -> list[str] | None:
    if not isinstance(value, str) or not value:
        return None
    result = list(dict.fromkeys(item.strip() for item in value.split(",") if item.strip()))
    return result or None


def get_service(session: AsyncSession = Depends(get_db)) -> ReportingService:
    return ReportingService(ReportingRepository(session))


def csv_response(filename: str, headers: list[str], rows: list[list[object]]) -> Response:
    output = StringIO()
    writer = csv.writer(output, lineterminator="\r\n")
    writer.writerow(headers)
    writer.writerows(rows)
    return Response(
        content="\ufeff" + output.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/chat-history", response_model=ChatHistoryResponse)
async def chat_history(
    from_date: date = Query(..., alias="from"),
    to_date: date = Query(..., alias="to"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current: AuthSession = Depends(require_authenticated_session),
    service: ReportingService = Depends(get_service),
):
    try:
        return await service.chat_histories(from_date, to_date, page, page_size, current)
    except ReportingError as error:
        raise HTTPException(status_code=422, detail=str(error)) from None


@router.get("/chat-history/export.xlsx")
async def chat_history_export(
    from_date: date = Query(..., alias="from"), to_date: date = Query(..., alias="to"),
    answer_type: str | None = Query(None, pattern="^(FAQ|GENERATED_AI)$"),
    rating: str | None = Query(None, pattern="^(RATED|GOOD|BAD|NONE)$"),
    comment: str | None = Query(None, pattern="^(WITH|WITHOUT)$"),
    role: str | None = Query(None, pattern="^(staff|admin)$"), user_ids: str | None = Query(None, max_length=5000),
    current: AuthSession = Depends(require_authenticated_session), service: ReportingService = Depends(get_service),
):
    try: start_at, end_at = utc_period(from_date, to_date)
    except ReportingError as error: raise HTTPException(status_code=422, detail=str(error)) from None
    own_key = None
    if current.role == "staff":
        from app.services.analytics_service import AnalyticsService
        own_key = AnalyticsService(service.repository).visitor_key("AUTHENTICATED", f"{current.site}:{current.subject}")
        role, user_ids = None, None
    rows = await service.repository.chat_history_export(start_at, end_at, visitor_key=own_key, answer_type=answer_type if isinstance(answer_type, str) else None, rating=rating if isinstance(rating, str) else None, comment=comment if isinstance(comment, str) else None, role=role if isinstance(role, str) else None, user_ids=parsed_user_ids(user_ids))
    workbook = Workbook(); sheet = workbook.active; sheet.title = "チャット履歴"
    headers = ["セッションID", "応答番号", "利用者ID", "利用者氏名", "ユーザ種別", "サイト", "質問日時", "回答日時", "回答種別", "質問", "回答", "評価", "コメント"]
    sheet.append(headers)
    for cell in sheet[1]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="1F4E78")
    excel_datetime = lambda value: value.astimezone().replace(tzinfo=None) if value else None
    for row in rows: sheet.append([str(row["session_id"]), row["sequence_number"], row.get("subject"), row.get("display_name"), row.get("role"), row.get("site"), excel_datetime(row["question_submitted_at"]), excel_datetime(row.get("answer_displayed_at")), row.get("answer_type"), row.get("question_text"), row.get("answer_text"), row.get("rating"), row.get("comment")])
    for column, width in {"A":38,"B":10,"C":18,"D":24,"E":16,"F":14,"G":22,"H":22,"I":16,"J":48,"K":64,"L":12,"M":36}.items(): sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"; output = BytesIO(); workbook.save(output)
    return Response(content=output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="history{datetime.now():%Y%m%d%H%M}.xlsx"'})


@router.get("/usage/users.csv")
async def usage_users_csv(
    from_date: date = Query(..., alias="from"),
    to_date: date = Query(..., alias="to"),
    role: str | None = Query(None, pattern="^(staff|admin)$"),
    _: AuthSession = Depends(require_system_admin_session),
    service: ReportingService = Depends(get_service),
):
    try:
        start_at, end_at = utc_period(from_date, to_date)
    except ReportingError as error:
        raise HTTPException(status_code=422, detail=str(error)) from None
    role = role if isinstance(role, str) else None
    rows = await service.repository.usage_users(start_at, end_at, role=role)
    return csv_response(
        f"usage_users_{from_date:%Y%m%d}_{to_date:%Y%m%d}.csv",
        ["利用者ID", "利用者氏名", "ロール", "サイト", "認証種別", "初回記録日時", "最終記録日時", "アクセス数", "チャット数"],
        [[
            row.get("subject") or f"利用者-{row['visitor_key'][:12]}",
            row.get("display_name") or "",
            row.get("role") or "",
            row.get("site") or "",
            row["identity_kind"], row["created_at"], row["last_seen_at"],
            row["access_count"], row["chat_count"],
        ] for row in rows],
    )


@router.get("/usage/access-logs.csv")
async def access_logs_csv(
    from_date: date = Query(..., alias="from"),
    to_date: date = Query(..., alias="to"),
    surface: str | None = Query(None, pattern="^(CHAT|ADMIN)$"),
    role: str | None = Query(None, pattern="^(staff|admin)$"),
    user_ids: str | None = Query(None, max_length=5000),
    _: AuthSession = Depends(require_system_admin_session),
    service: ReportingService = Depends(get_service),
):
    try:
        start_at, end_at = utc_period(from_date, to_date)
    except ReportingError as error:
        raise HTTPException(status_code=422, detail=str(error)) from None
    surface = surface if isinstance(surface, str) else None
    role = role if isinstance(role, str) else None
    rows = await service.repository.access_logs(
        start_at, end_at, surface=surface, role=role, user_ids=parsed_user_ids(user_ids)
    )
    return csv_response(
        f"access_logs_{from_date:%Y%m%d}_{to_date:%Y%m%d}.csv",
        ["アクセスID", "利用者ID", "利用者氏名", "ロール", "サイト", "利用画面", "認証種別", "アクセス日時", "記録日時"],
        [[
            row["id"], row.get("subject") or f"利用者-{row['visitor_key'][:12]}",
            row.get("display_name") or "", row.get("role") or "", row.get("site") or "", row.get("surface") or "",
            row["identity_kind"], row["accessed_at"], row["recorded_at"],
        ] for row in rows],
    )


@router.get("/usage/operation-logs.csv")
async def operation_logs_csv(
    from_date: date = Query(..., alias="from"),
    to_date: date = Query(..., alias="to"),
    surface: str | None = Query(None, pattern="^(CHAT|ADMIN)$"),
    operation_type: str | None = Query(None, pattern="^(CREATE|UPDATE|DELETE|DOWNLOAD|UPLOAD)$"),
    role: str | None = Query(None, pattern="^(staff|admin)$"),
    user_ids: str | None = Query(None, max_length=5000),
    _: AuthSession = Depends(require_system_admin_session),
    service: ReportingService = Depends(get_service),
):
    try:
        start_at, end_at = utc_period(from_date, to_date)
    except ReportingError as error:
        raise HTTPException(status_code=422, detail=str(error)) from None
    surface = surface if isinstance(surface, str) else None
    operation_type = operation_type if isinstance(operation_type, str) else None
    role = role if isinstance(role, str) else None
    rows = await service.repository.operation_logs(
        start_at, end_at, role=role, user_ids=parsed_user_ids(user_ids)
    )
    if surface == "CHAT":
        rows = [row for row in rows if row["request_path"].startswith("/api/v1/chat/")]
    elif surface == "ADMIN":
        rows = [row for row in rows if not row["request_path"].startswith("/api/v1/chat/")]
    if operation_type:
        rows = [row for row in rows if operation_kind(row["http_method"], row["request_path"]) == operation_type]
    return csv_response(
        f"operation_logs_{from_date:%Y%m%d}_{to_date:%Y%m%d}.csv",
        ["操作ID", "操作者ID", "操作者氏名", "ロール", "サイト", "操作内容", "HTTPメソッド", "操作先", "結果コード", "操作日時"],
        [[
            row["id"], row.get("operator_subject") or f"利用者-{row['operator_key'][:12]}",
            row.get("operator_display_name") or "", row["operator_role"], row.get("operator_site") or "",
            operation_description(row["http_method"], row["request_path"]),
            row["http_method"], row["request_path"], row["status_code"], row["operated_at"],
        ] for row in rows],
    )
