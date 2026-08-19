from datetime import datetime, timezone
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import AsyncMock

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from pydantic import ValidationError

from app.api.v1.faqs import get_service
from app.main import app
from app.repositories.faq import FaqRepository
from app.schemas.faq import FaqBulkDeleteRequest, FaqCreateRequest, FaqDeleteTarget, FaqFilters
from app.services.faq_service import FaqError, FaqService


def make_assignment(index: int, value_name: str):
    type_row = SimpleNamespace(id=index, type_code=f"FAQ_TYPE_{index}", display_label=f"区分{index}", display_order=index)
    value_row = SimpleNamespace(id=index * 10, value_name=value_name)
    return SimpleNamespace(
        classification_type_id=index, classification_value_id=index * 10,
        classification_type=type_row, classification_value=value_row,
    )


def make_faq(faq_id: int = 1, *, question: str = "申請期限は？", answer: str = "8月末です。", chat_enabled: bool = True, version: int = 1):
    return SimpleNamespace(
        id=faq_id, question=question, answer=answer, chat_enabled=chat_enabled, version=version,
        created_at=datetime(2026, 8, 1, tzinfo=timezone.utc),
        updated_at=datetime(2026, 8, faq_id, 1, 2, tzinfo=timezone.utc),
        similar_questions=[],
        classification_assignments=[make_assignment(1, "奨学金"), make_assignment(4, "神楽坂")],
    )


def create_payload(**values):
    return FaqCreateRequest(
        question=values.get("question", "質問"), answer=values.get("answer", "回答"),
        similar_questions=values.get("similar_questions", []),
        classification_1_value_id=values.get("classification_1_value_id"),
        classification_2_value_id=values.get("classification_2_value_id"),
        classification_3_value_id=values.get("classification_3_value_id"),
        classification_4_value_id=values.get("classification_4_value_id"),
        chat_enabled=values.get("chat_enabled", True),
    )


def create_repository(detail=None):
    repository = AsyncMock()
    repository.create.return_value = 1
    repository.get_detail.return_value = detail or make_faq()
    return repository


@pytest.fixture
def mock_service():
    service = AsyncMock()
    app.dependency_overrides[get_service] = lambda: service
    yield service
    app.dependency_overrides.clear()


def test_faq_filters_defaults_and_validation():
    filters = FaqFilters(keyword=" 期限 ")
    assert (filters.keyword, filters.sort, filters.order, filters.page, filters.page_size) == ("期限", "updated_at", "desc", 1, 10)
    assert FaqFilters(keyword="  ").keyword is None
    for size in (10, 20, 50, 100):
        assert FaqFilters(page_size=size).page_size == size
    for values in ({"sort": "question"}, {"order": "sideways"}, {"page_size": 5}):
        with pytest.raises(ValidationError):
            FaqFilters(**values)


def test_keyword_or_and_all_other_filters_are_and_conditions():
    filters = FaqFilters(
        keyword="申請", classification_1_value_id=10, classification_2_value_id=20,
        classification_3_value_id=30, classification_4_value_id=40, chat_enabled=True,
    )
    conditions = FaqRepository.conditions(filters, {f"FAQ_TYPE_{i}": i for i in range(1, 5)})
    assert len(conditions) == 6
    assert "faqs.question" in str(conditions[0]) and "faqs.answer" in str(conditions[0]) and " OR " in str(conditions[0])
    assert all("faq_similar_questions" not in str(condition) for condition in conditions)


@pytest.mark.anyio
async def test_list_serializes_dynamic_classifications_and_paging():
    repository = AsyncMock()
    repository.list.return_value = ([make_faq(2), make_faq(1)], 12, 2)
    result = await FaqService(repository).list(FaqFilters())
    assert result.total_count == 12 and result.total_pages == 2
    assert result.items[0].classifications[0].display_label == "区分1"
    assert result.items[0].classifications[1].value_name == "神楽坂"
    assert {item.type_code for item in result.items[0].classifications} == {"FAQ_TYPE_1", "FAQ_TYPE_4"}


@pytest.mark.anyio
async def test_zero_rows_and_page_not_found():
    repository = AsyncMock()
    repository.list.side_effect = [([], 0, 0), ([], 11, 2)]
    service = FaqService(repository)
    assert (await service.list(FaqFilters())).items == []
    with pytest.raises(FaqError, match="ページがありません") as error:
        await service.list(FaqFilters(page=3))
    assert error.value.code == "PAGE_NOT_FOUND"


@pytest.mark.anyio
async def test_each_classification_filter_is_validated_against_its_type():
    repository = AsyncMock()
    repository.resolve_value_type.side_effect = [1, 2, 3, 4]
    repository.list.return_value = ([], 0, 0)
    filters = FaqFilters(
        classification_1_value_id=10, classification_2_value_id=20,
        classification_3_value_id=30, classification_4_value_id=40,
    )
    await FaqService(repository).list(filters)
    assert repository.resolve_value_type.await_args_list[0].args == ("FAQ_TYPE_1", 10)
    assert repository.resolve_value_type.await_args_list[3].args == ("FAQ_TYPE_4", 40)


@pytest.mark.anyio
async def test_wrong_classification_value_is_422_domain_error():
    repository = AsyncMock()
    repository.resolve_value_type.return_value = None
    with pytest.raises(FaqError) as error:
        await FaqService(repository).list(FaqFilters(classification_1_value_id=999))
    assert error.value.code == "INVALID_FAQ_CLASSIFICATION"


@pytest.mark.anyio
async def test_single_delete_not_found_and_version_conflict():
    repository = AsyncMock()
    repository.get.side_effect = [None, make_faq()]
    repository.delete_one.return_value = False
    service = FaqService(repository)
    with pytest.raises(FaqError) as missing:
        await service.delete(99, 1)
    assert missing.value.code == "FAQ_NOT_FOUND"
    with pytest.raises(FaqError) as conflict:
        await service.delete(1, 9)
    assert conflict.value.code == "FAQ_VERSION_CONFLICT"


@pytest.mark.anyio
async def test_single_and_bulk_delete_delegate_atomic_repository_operations():
    repository = AsyncMock()
    repository.get.return_value = make_faq()
    repository.delete_one.return_value = True
    repository.bulk_delete.return_value = 2
    service = FaqService(repository)
    await service.delete(1, 1)
    payload = FaqBulkDeleteRequest(items=[FaqDeleteTarget(id=1, version=1), FaqDeleteTarget(id=2, version=2)])
    assert await service.bulk_delete(payload) == 2
    repository.bulk_delete.assert_awaited_once_with(payload.items)


@pytest.mark.anyio
@pytest.mark.parametrize("failure,code", [(LookupError("not_found"), "FAQ_NOT_FOUND"), (ValueError("version_mismatch"), "FAQ_VERSION_CONFLICT")])
async def test_bulk_delete_failure_is_reported_without_partial_success(failure, code):
    repository = AsyncMock()
    repository.bulk_delete.side_effect = failure
    payload = FaqBulkDeleteRequest(items=[FaqDeleteTarget(id=1, version=1)])
    with pytest.raises(FaqError) as error:
        await FaqService(repository).bulk_delete(payload)
    assert error.value.code == code


@pytest.mark.anyio
async def test_excel_uses_dynamic_labels_japanese_values_and_jst():
    repository = AsyncMock()
    repository.list.return_value = ([make_faq(chat_enabled=True), make_faq(2, chat_enabled=False)], 2, 1)
    content = await FaqService(repository).export_excel(FaqFilters(), {f"FAQ_TYPE_{i}": f"表示区分{i}" for i in range(1, 5)})
    sheet = load_workbook(BytesIO(content)).active
    assert [cell.value for cell in sheet[1]] == ["ID", "質問", "回答", "表示区分1", "表示区分2", "表示区分3", "表示区分4", "チャット利用", "更新日時"]
    assert sheet.cell(2, 4).value == "奨学金" and sheet.cell(2, 5).value is None
    assert sheet.cell(2, 8).value == "公開" and sheet.cell(3, 8).value == "非公開"
    assert sheet.cell(2, 9).value == "2026/08/01 10:02"


@pytest.mark.anyio
async def test_api_page_not_found_and_invalid_classification_codes(mock_service):
    for code in ("PAGE_NOT_FOUND", "INVALID_FAQ_CLASSIFICATION"):
        mock_service.list.side_effect = FaqError(code, "エラー")
        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
            response = await client.get("/api/v1/faqs")
        assert response.status_code == 422
        assert response.json()["detail"]["code"] == code


@pytest.mark.anyio
async def test_api_delete_codes_and_empty_bulk_delete(mock_service):
    mock_service.delete.side_effect = FaqError("FAQ_VERSION_CONFLICT", "競合")
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        conflict = await client.delete("/api/v1/faqs/1?version=1")
        empty = await client.post("/api/v1/faqs/bulk-delete", json={"items": []})
    assert conflict.status_code == 409 and conflict.json()["detail"]["code"] == "FAQ_VERSION_CONFLICT"
    assert empty.status_code == 422
    mock_service.bulk_delete.assert_not_awaited()


@pytest.mark.anyio
async def test_export_filename(mock_service):
    mock_service.repository.list_type_labels.return_value = {}
    mock_service.export_excel.return_value = b"xlsx"
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        response = await client.get("/api/v1/faqs/export")
    assert response.status_code == 200
    assert "faq" in response.headers["content-disposition"] and response.headers["content-disposition"].endswith(".xlsx")


def test_repository_uses_fixed_count_eager_loading_not_per_row_queries():
    source = FaqRepository.list.__code__.co_names
    assert "selectinload" in source


@pytest.mark.anyio
async def test_minimal_faq_create_trims_and_returns_complete_detail():
    repository = create_repository(make_faq(chat_enabled=False))
    result = await FaqService(repository).create(create_payload(question=" 質問 ", answer=" 回答\n本文 ", chat_enabled=False))
    assert result.id == 1 and result.version == 1 and result.created_at
    assert result.similar_questions == [] and result.chat_enabled is False
    repository.create.assert_awaited_once_with(
        question="質問", answer="回答\n本文", similar_questions=[], classifications=[], chat_enabled=False,
    )
    repository.commit.assert_awaited_once()


@pytest.mark.anyio
async def test_create_multiple_similar_questions_preserves_order_and_all_classifications():
    detail = make_faq()
    detail.similar_questions = [
        SimpleNamespace(id=11, question="類似1", display_order=1),
        SimpleNamespace(id=12, question="類似2", display_order=2),
    ]
    detail.classification_assignments = [make_assignment(index, f"値{index}") for index in range(1, 5)]
    repository = create_repository(detail)
    repository.get_value_type.side_effect = [(1, "FAQ_TYPE_1"), (2, "FAQ_TYPE_2"), (3, "FAQ_TYPE_3"), (4, "FAQ_TYPE_4")]
    result = await FaqService(repository).create(create_payload(
        similar_questions=[" 類似1 ", "類似2"],
        classification_1_value_id=10, classification_2_value_id=20,
        classification_3_value_id=30, classification_4_value_id=40,
    ))
    assert [(item.question, item.display_order) for item in result.similar_questions] == [("類似1", 1), ("類似2", 2)]
    assert len(result.classifications) == 4
    assert result.classifications[0].display_label == "区分1" and result.classifications[0].value_name == "値1"
    assert repository.create.await_args.kwargs["classifications"] == [(1, 10), (2, 20), (3, 30), (4, 40)]


@pytest.mark.anyio
async def test_create_one_similar_question_and_one_classification():
    detail = make_faq()
    detail.similar_questions = [SimpleNamespace(id=11, question="類似1", display_order=1)]
    detail.classification_assignments = [make_assignment(1, "値1")]
    repository = create_repository(detail)
    repository.get_value_type.return_value = (1, "FAQ_TYPE_1")
    result = await FaqService(repository).create(create_payload(
        similar_questions=[" 類似1 "], classification_1_value_id=10,
    ))
    assert [(item.question, item.display_order) for item in result.similar_questions] == [("類似1", 1)]
    assert repository.create.await_args.kwargs["classifications"] == [(1, 10)]


@pytest.mark.anyio
@pytest.mark.parametrize("field,value,code", [
    ("question", "", "FAQ_QUESTION_REQUIRED"), ("question", "   ", "FAQ_QUESTION_REQUIRED"),
    ("question", "x" * 501, "FAQ_QUESTION_TOO_LONG"),
    ("answer", "", "FAQ_ANSWER_REQUIRED"), ("answer", "   ", "FAQ_ANSWER_REQUIRED"),
    ("answer", "x" * 1001, "FAQ_ANSWER_TOO_LONG"),
])
async def test_question_and_answer_validation(field, value, code):
    repository = create_repository()
    with pytest.raises(FaqError) as error:
        await FaqService(repository).create(create_payload(**{field: value}))
    assert error.value.code == code
    repository.create.assert_not_awaited()
    repository.rollback.assert_awaited_once()


@pytest.mark.anyio
@pytest.mark.parametrize("question,answer", [("x", "y"), ("x" * 500, "y" * 1000)])
async def test_question_and_answer_boundaries_are_allowed(question, answer):
    repository = create_repository()
    await FaqService(repository).create(create_payload(question=question, answer=answer))
    assert repository.create.await_args.kwargs["question"] == question
    assert repository.create.await_args.kwargs["answer"] == answer


@pytest.mark.anyio
@pytest.mark.parametrize("value,code", [
    ("", "FAQ_SIMILAR_QUESTION_REQUIRED"), ("   ", "FAQ_SIMILAR_QUESTION_REQUIRED"),
    ("x" * 501, "FAQ_SIMILAR_QUESTION_TOO_LONG"),
])
async def test_similar_question_validation(value, code):
    repository = create_repository()
    with pytest.raises(FaqError) as error:
        await FaqService(repository).create(create_payload(similar_questions=["正常", value]))
    assert error.value.code == code
    repository.create.assert_not_awaited()


@pytest.mark.anyio
async def test_similar_question_500_chars_and_duplicates_are_allowed():
    repository = create_repository()
    value = "x" * 500
    await FaqService(repository).create(create_payload(similar_questions=[value, value]))
    assert repository.create.await_args.kwargs["similar_questions"] == [value, value]


@pytest.mark.anyio
@pytest.mark.parametrize("resolved,code", [
    (None, "FAQ_CLASSIFICATION_NOT_FOUND"), ((2, "FAQ_TYPE_2"), "INVALID_FAQ_CLASSIFICATION"),
])
async def test_create_classification_must_exist_and_belong_to_expected_type(resolved, code):
    repository = create_repository()
    repository.get_value_type.return_value = resolved
    with pytest.raises(FaqError) as error:
        await FaqService(repository).create(create_payload(classification_1_value_id=20))
    assert error.value.code == code
    repository.create.assert_not_awaited()


@pytest.mark.anyio
@pytest.mark.parametrize("failure", ["similar question insert failed", "assignment insert failed"])
async def test_create_failure_rolls_back_without_commit(failure):
    repository = create_repository()
    repository.create.side_effect = RuntimeError(failure)
    with pytest.raises(RuntimeError):
        await FaqService(repository).create(create_payload(similar_questions=["類似"] ))
    repository.rollback.assert_awaited_once()
    repository.commit.assert_not_awaited()


@pytest.mark.anyio
async def test_create_api_returns_201_and_domain_validation_code(mock_service):
    detail = FaqService.serialize_detail(make_faq())
    mock_service.create.side_effect = [detail, FaqError("FAQ_QUESTION_REQUIRED", "質問を入力してください。")]
    payload = {"question": "質問", "answer": "回答", "similar_questions": [], "chat_enabled": True}
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        created = await client.post("/api/v1/faqs", json=payload)
        invalid = await client.post("/api/v1/faqs", json=payload)
    assert created.status_code == 201 and created.json()["version"] == 1
    assert invalid.status_code == 422 and invalid.json()["detail"]["code"] == "FAQ_QUESTION_REQUIRED"


@pytest.mark.anyio
async def test_detail_api_returns_404_code(mock_service):
    mock_service.get_detail.side_effect = FaqError("FAQ_NOT_FOUND", "指定されたFAQが見つかりません。")
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        response = await client.get("/api/v1/faqs/999")
    assert response.status_code == 404 and response.json()["detail"]["code"] == "FAQ_NOT_FOUND"
